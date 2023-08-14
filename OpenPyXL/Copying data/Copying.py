#!/usr/bin/python

# -*- coding: utf-8 -*-

"""
Could you please suggest how to copy the data from on work book to other book with specified rows
Source: Excel work book "WB1" having work sheet "WS1", This sheet having 1000 rows of data
Destination: New work book 'WB2' and work sheets WS1,WS2...WS10
Could you please suggest the code for following condition:
Copy the first 100 rows data and paste it WS1 sheet
Copy the next 200 rows data and paste it WS2 sheet
Copy the next 50 rows data and paste it WS3 sheet
Copy the next 300 rows data and paste it WS4 sheet
Copy the next 350 rows data and paste it WS4 sheet
"""

from openpyxl import Workbook, load_workbook

WB1 = load_workbook("Source.xlsx", data_only=True)
WB1_WS1 = WB1["WS1"]
WB2 = Workbook()

# Create WB2 sheets WS1-WS10
for i in range(1, 11):
    WB2.create_sheet(f"WS{i}")

    # delete first sheet
    WB2.remove(WB2.worksheets[0])

# Define the copy ranges and sheets
copy_ranges = [100, 200, 50, 300, 350]
copy_to_sheets = ["WS1", "WS2", "WS3", "WS4", "WS4"]

# Copy the values from the rows in WB1 to WB2.
for i in range(len(copy_ranges)):
    # Set the sheet to copy to
    ws = WB2[copy_to_sheets[i]]
# Initialize row offset
offset = 1
# Set the row offset
for s in range(i):
    offset += copy_ranges[s]

# Copy the row with the help of iter_rows, append the row
for j in range(offset, offset + copy_ranges[i]):
    # if j == 0:
    # continue
    for row in WB1_WS1.iter_rows(
        min_row=j, max_row=j, min_col=1, max_col=WB1_WS1.max_column
    ):
        values_row = [cell.value for cell in row]
        ws.append(values_row)

# Save the workbook
WB2.save("WB2.xlsx")
