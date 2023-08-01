#!/usr/bin/env python3

"""
Inserting rows into a table for existing workbook
"""
from openpyxl import load_workbook

SAVE_FILENAME = "New Table.xlsx"

wb = load_workbook("Table.xlsx")
ws = wb.active

# Insert rows into table
ws.insert_rows(idx=5, amount=1)

table = ws.tables["Table1"]
table.ref = "A2:B8"
ws["A5"] = 6
ws["B5"] = "Cucumber"

wb.save(SAVE_FILENAME)
