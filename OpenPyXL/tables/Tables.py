from openpyxl import load_workbook

filename = "Tables.xlsx"

wb = load_workbook(filename)

ws = wb.worksheets[0]

ws_tables = []

ws["E6"] = "VW"
ws["F6"] = "Polo"
ws["A6"] = "Peter"
ws["B6"] = "Dinkle"
ws["C6"] = "USA"


for table in ws._tables:
    ws_tables.append(table)
    if table.name == "Cars":
        table.ref = "E2:F6"
    if table.name == "Customers":
        table.ref = "A2:C6"
    print(table.name, table.ref)

wb.save(filename)
