#!/usr/bin/env python3

"""
Creating an xml invoice from Excel
"""
from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook("invoice excel.xlsx", data_only=True)
ws = wb.active

date = ws["B1"].value
customer_nr = ws["B2"].value

invoice_row_list: list = []

for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=5):
    row = [cell.value for cell in row]
    invoice_row_list.append(row)

# print(invoice_row_list)

"""
<Invoice>
    Date
    Customer number
    <Invoice row>
        Name
        Amount
        Unit
        Price
        Total
    </Invoice row>
</Invoice>
"""

doc, tag, text = Doc().tagtext()

with tag("Invoice"):
    with tag("Date"):
        text(str(date))
    with tag("Customer number"):
        text(customer_nr)
    for row in invoice_row_list:
        with tag("Invoice row"):
            with tag("Name"):
                text(row[0])
            with tag("Amount"):
                text(row[1])
            with tag("Unit"):
                text(row[2])
            with tag("Price"):
                text(row[3])
            with tag("Total"):
                text(row[4])

result = indent(doc.getvalue(), indentation="    ", indent_text=False)

print(result)

with open("Generated invoice.xml", "w", encoding="utf-8") as f:
    f.write(result)
